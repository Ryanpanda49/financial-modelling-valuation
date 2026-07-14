import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fmva import ValuationModel
from fmva.data.models import SelectionMethod
from fmva.data.statement_builder import HistoricalStatements

FIXTURE = Path("data/sample/cost_fy2021_2025_history.json")


def test_public_cost_fixture_runs_complete_offline_model() -> None:
    payload = json.loads(FIXTURE.read_text(encoding="utf-8"))
    history = HistoricalStatements.from_dict(payload)
    parent_income = next(
        item
        for item in history.observations
        if item.account == "net_income_attributable_to_parent"
        and item.fiscal_year == 2025
    )
    model = ValuationModel.from_history_json(
        FIXTURE,
        forecast_assumptions_path="examples/cost/forecast_assumptions.yaml",
        valuation_assumptions_path="examples/cost/valuation_assumptions.yaml",
    )
    result = model.run()

    assert payload["metadata"]["fiscal_years"] == [2021, 2022, 2023, 2024, 2025]
    assert len(history.observations) == 330
    assert not history.quality_issues
    assert parent_income.provenance.selection_method is SelectionMethod.DERIVED
    assert any("minority_interest" in warning for warning in parent_income.provenance.warnings)
    assert list(result.forecast.income_statement.columns) == [2026, 2027, 2028, 2029, 2030]
    assert all(check.status.value == "PASS" for check in result.historical_checks)
    assert all(check.status.value == "PASS" for check in result.forecast.checks)
    assert all(check.status.value == "PASS" for check in result.dcf.checks)
    assert result.dcf.implied_share_price == pytest.approx(302.27, abs=0.02)


def test_cost_bottom_up_drivers_flow_through_statements_dcf_and_checks(tmp_path: Path) -> None:
    model = ValuationModel.from_history_json(
        FIXTURE,
        forecast_assumptions_path="examples/cost/forecast_assumptions.yaml",
        valuation_assumptions_path="examples/cost/valuation_assumptions.yaml",
        business_driver_path="examples/cost/business_drivers.yaml",
    )

    result = model.run()
    assert result.forecast.business_drivers is not None
    assert result.forecast.income_statement.loc["revenue", 2026] == pytest.approx(
        result.forecast.business_drivers.loc["total_revenue", 2026]
    )
    assert all(check.status.value == "PASS" for check in result.forecast.checks)
    assert sum(
        check.check == "business_driver_revenue_tie" for check in result.forecast.checks
    ) == 5
    assert result.dcf.implied_share_price > 0

    tables = result.export_tables(tmp_path / "tables")
    workbook = result.export_excel(tmp_path / "cost_bottom_up.xlsx")
    report = result.export_markdown(tmp_path / "cost_bottom_up.md")
    assert "business_drivers" in tables
    assert workbook.exists()
    assert "Business_Drivers" in load_workbook(workbook, read_only=True).sheetnames
    assert "## Business Driver Model" in report.read_text(encoding="utf-8")
