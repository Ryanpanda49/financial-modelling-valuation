from pathlib import Path

from openpyxl import load_workbook

from fmva.analysis.ratios import calculate_financial_ratios
from fmva.forecasting.assumptions import ForecastAssumptions
from fmva.forecasting.three_statement import InitialFinancialState, ThreeStatementModel
from fmva.output import ModelResult
from fmva.valuation.dcf import value_dcf
from fmva.valuation.models import ValuationAssumptions
from fmva.valuation.sensitivity import wacc_terminal_growth_sensitivity


def _model_result() -> ModelResult:
    state = InitialFinancialState(
        fiscal_year=2024,
        revenue=1000.0,
        cash_and_equivalents=100.0,
        accounts_receivable=100.0,
        inventory=80.0,
        other_current_assets=20.0,
        property_plant_equipment=300.0,
        other_assets=400.0,
        accounts_payable=70.0,
        accrued_liabilities=30.0,
        short_term_debt=50.0,
        long_term_debt=150.0,
        other_liabilities=200.0,
        contributed_equity=250.0,
        retained_earnings=250.0,
    )
    assumptions = ForecastAssumptions.from_yaml(
        "config/forecast_assumptions.example.yaml"
    )
    valuation = ValuationAssumptions.from_yaml(
        "config/valuation_assumptions.example.yaml"
    )
    forecast = ThreeStatementModel().run(state, assumptions)
    ratios = calculate_financial_ratios(forecast, state)
    dcf = value_dcf(forecast, valuation)
    sensitivity = wacc_terminal_growth_sensitivity(
        forecast,
        valuation,
        [0.07, 0.08, 0.09],
        [0.01, 0.02, 0.03],
    )
    return ModelResult(
        company_name="Example Company",
        ticker="EXM",
        as_of="2024-12-31",
        forecast=forecast,
        ratios=ratios,
        dcf=dcf,
        sensitivity=sensitivity,
        assumption_summary=[
            {"Assumption": "Revenue growth", "2025": "5.0%", "Source": "User input"},
            {"Assumption": "Tax rate", "2025": "25.0%", "Source": "User input"},
        ],
        limitations=("Synthetic opening balance sheet; not an investment recommendation.",),
    )


def test_model_result_exports_markdown_and_structured_tables(tmp_path: Path) -> None:
    result = _model_result()
    report = result.export_markdown(tmp_path / "report.md")
    content = report.read_text(encoding="utf-8")

    assert "## Projected Financials" in content
    assert "## DCF Valuation" in content
    assert "## Sensitivity Analysis" in content
    assert "## Model Checks" in content
    assert "does not constitute investment advice" in content

    tables = result.export_tables(tmp_path / "tables")
    assert set(tables) == {
        "income_statement",
        "balance_sheet",
        "cash_flow_statement",
        "working_capital",
        "fixed_assets",
        "debt_schedule",
        "financial_ratios",
        "dcf_forecast",
        "sensitivity",
    }
    assert all(path.exists() and path.stat().st_size > 100 for path in tables.values())


def test_model_result_exports_six_valid_png_charts(tmp_path: Path) -> None:
    charts = _model_result().export_charts(tmp_path / "charts")

    assert set(charts) == {
        "revenue_trend",
        "margin_trend",
        "net_income",
        "cfo_and_fcf",
        "cash_and_debt",
        "dcf_sensitivity",
    }
    for path in charts.values():
        assert path.stat().st_size > 1000
        assert path.read_bytes()[:8] == b"\x89PNG\r\n\x1a\n"


def test_model_result_exports_complete_blue_excel_workbook(tmp_path: Path) -> None:
    path = _model_result().export_excel(tmp_path / "model.xlsx")
    workbook = load_workbook(path, data_only=False)

    assert workbook.sheetnames == [
        "Summary",
        "Sources_Audit",
        "Historical",
        "Assumptions",
        "Income_Statement",
        "Balance_Sheet",
        "Cash_Flow",
        "Working_Capital",
        "CapEx_Dep",
        "Debt_Schedule",
        "Ratios",
        "DCF",
        "Sensitivity",
        "Checks",
    ]
    assert workbook["Summary"]["A1"].fill.fgColor.rgb.endswith("17365D")
    assert workbook["Assumptions"]["B5"].font.color.rgb.endswith("0000FF")
    assert workbook["Summary"]["B9"].value.startswith("='DCF'!")
    assert "SUMPRODUCT" in workbook["Sensitivity"]["B5"].value
    assert len(workbook["Summary"]._charts) == 1
    assert workbook["Summary"]._charts[0].series[0].tx.v == "Revenue"
    assert workbook["Summary"]._charts[0].series[1].tx.v == "EBITDA"
    assert not workbook._external_links

    formula_errors = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if any(error in cell.value for error in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}")
    assert formula_errors == []
