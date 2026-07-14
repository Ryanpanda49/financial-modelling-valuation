"""Public, reproducible Excel exporter with an analyst-style blue theme."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fmva.output.result_types import ModelResultData

NAVY = "17365D"
PRIMARY_BLUE = "1F4E78"
MEDIUM_BLUE = "5B9BD5"
ACCENT_BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
INPUT_BLUE = "0000FF"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN = "008000"
RED = "C00000"
AMBER = "FFF2CC"
LIGHT_GRAY = "E7E6E6"
THIN_BLUE = Side(style="thin", color="9EBCD4")
MEDIUM_BORDER = Side(style="medium", color=PRIMARY_BLUE)
FINANCIAL_FORMAT = '#,##0.0;[Red](#,##0.0);-'
PERCENT_FORMAT = '0.0%;[Red](0.0%);-'
MULTIPLE_FORMAT = '0.0x;[Red](0.0x);-'
PRICE_FORMAT = '$0.00;[Red]($0.00);-'


def export_excel(data: ModelResultData, path: str | Path) -> Path:
    """Export the complete public workbook contract as a styled XLSX file."""

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_names = (
        "Summary",
        "Sources_Audit",
        "Historical",
        "Assumptions",
        "Income_Statement",
        "Balance_Sheet",
        "Cash_Flow",
        "Working_Capital",
        "CapEx_Dep",
        "Debt_Schedule",
        "Ratios",
        "DCF",
        "Sensitivity",
        "Checks",
    )
    sheets = {name: workbook.create_sheet(name) for name in sheet_names}
    for sheet in sheets.values():
        _prepare_sheet(sheet)

    _write_sources(sheets["Sources_Audit"], data)
    _write_historical(sheets["Historical"], data)
    _write_assumptions(sheets["Assumptions"], data.assumption_summary)
    _write_financial_table(
        sheets["Income_Statement"], "Projected Income Statement", data.forecast.income_statement
    )
    _write_financial_table(
        sheets["Balance_Sheet"], "Projected Balance Sheet", data.forecast.balance_sheet
    )
    _write_financial_table(
        sheets["Cash_Flow"], "Projected Cash Flow Statement", data.forecast.cash_flow_statement
    )
    _write_financial_table(
        sheets["Working_Capital"], "Working Capital Schedule", data.forecast.working_capital
    )
    _write_financial_table(
        sheets["CapEx_Dep"], "CapEx and Depreciation Schedule", data.forecast.fixed_assets
    )
    _write_financial_table(
        sheets["Debt_Schedule"], "Debt and Interest Schedule", data.forecast.debt_schedule
    )
    _write_ratio_sheet(sheets["Ratios"], data)
    dcf_cells = _write_dcf_sheet(sheets["DCF"], data)
    _write_sensitivity_sheet(sheets["Sensitivity"], data, dcf_cells)
    check_last_row = _write_checks_sheet(sheets["Checks"], data)
    _write_summary(sheets["Summary"], data, dcf_cells, check_last_row)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(target)
    return target


def _prepare_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B5"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True


def _title(sheet: Worksheet, title: str, last_column: int) -> None:
    last = max(last_column, 6)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last)
    cell = sheet.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28


def _table_header(sheet: Worksheet, row: int, labels: Iterable[object]) -> None:
    values = list(labels)
    for column, value in enumerate(values, 1):
        cell = sheet.cell(row, column, value)
        cell.fill = PatternFill("solid", fgColor=PRIMARY_BLUE)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="right" if column > 1 else "left")
        cell.border = Border(bottom=MEDIUM_BORDER)


def _write_financial_table(sheet: Worksheet, title: str, frame: pd.DataFrame) -> None:
    years = list(frame.columns)
    _title(sheet, title, len(years) + 1)
    sheet["A2"] = "USD millions unless otherwise indicated | Forecast values"
    sheet["A2"].font = Font(italic=True, color=PRIMARY_BLUE)
    _table_header(sheet, 4, ["Account", *years])
    for row, (label, values) in enumerate(frame.iterrows(), 5):
        label_cell = sheet.cell(row, 1, _display_label(str(label)))
        label_cell.alignment = Alignment(indent=1 if _is_detail_row(str(label)) else 0)
        for column, value in enumerate(values, 2):
            cell = sheet.cell(row, column, _excel_value(value))
            cell.number_format = _number_format_for(str(label), value)
            cell.alignment = Alignment(horizontal="right")
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        if _is_total_row(str(label)):
            for column in range(1, len(years) + 2):
                cell = sheet.cell(row, column)
                cell.font = Font(bold=True, color=BLACK)
                cell.border = Border(top=THIN_BLUE)
    _size_financial_sheet(sheet, len(years) + 1)


def _write_assumptions(sheet: Worksheet, records: list[dict[str, object]]) -> None:
    columns = _ordered_record_columns(records)
    _title(sheet, "Forecast Assumptions", len(columns))
    sheet["A2"] = "Blue font indicates researcher-editable inputs. Values originate in user configuration."
    sheet["A2"].font = Font(italic=True, color=PRIMARY_BLUE)
    if not columns:
        sheet["A4"] = "No assumptions supplied."
        return
    _table_header(sheet, 4, columns)
    for row, record in enumerate(records, 5):
        for column, key in enumerate(columns, 1):
            value = record.get(key)
            display_value = _display_label(str(value)) if key == "Assumption" else _excel_value(value)
            cell = sheet.cell(row, column, display_value)
            if key not in {"Assumption", "Source"}:
                linked_history = (
                    key == "Value"
                    and record.get("Source") == "Standardized historical bridge"
                )
                cell.font = Font(color=GREEN if linked_history else INPUT_BLUE)
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                cell.comment = Comment(
                    "Linked from standardized historical data; update through the source adapter."
                    if linked_history
                    else "Editable assumption from user configuration.",
                    "FMVA",
                )
                if key == "Value":
                    unit = str(record.get("Unit", ""))
                    cell.number_format = (
                        PERCENT_FORMAT
                        if unit == "Percent"
                        else MULTIPLE_FORMAT
                        if unit == "Multiple"
                        else FINANCIAL_FORMAT
                    )
                else:
                    cell.number_format = _assumption_number_format(
                        str(record.get("Assumption", ""))
                    )
            elif key == "Source":
                cell.font = Font(color=PRIMARY_BLUE, italic=True)
    sheet.freeze_panes = "B5"
    sheet.column_dimensions["A"].width = 40
    for column, key in enumerate(columns[1:], 2):
        if key in {"Source", "Rationale"}:
            width = 36
        elif key == "Source URL":
            width = 48
        elif key == "Unit":
            width = 20
        elif key == "Value":
            width = 18
        else:
            width = 14
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_ratio_sheet(sheet: Worksheet, data: ModelResultData) -> None:
    _title(sheet, "Financial Ratios", max(len(data.ratios.table.columns) + 1, 6))
    row = 4
    if data.historical_ratios is not None:
        row = _write_ratio_block(sheet, row, "Historical Ratios", data.historical_ratios.table)
        row += 2
    row = _write_ratio_block(sheet, row, "Forecast Ratios", data.ratios.table)
    warning_row = row + 2
    sheet.cell(warning_row, 1, "Ratio warnings")
    sheet.cell(warning_row, 1).fill = PatternFill("solid", fgColor=PRIMARY_BLUE)
    sheet.cell(warning_row, 1).font = Font(bold=True, color=WHITE)
    warnings = [
        *(data.historical_ratios.warnings if data.historical_ratios is not None else ()),
        *data.ratios.warnings,
    ]
    for offset, warning in enumerate(warnings, 1):
        sheet.cell(warning_row + offset, 1, warning)
        sheet.merge_cells(
            start_row=warning_row + offset,
            start_column=1,
            end_row=warning_row + offset,
            end_column=max(6, len(data.ratios.table.columns) + 1),
        )
    _size_financial_sheet(sheet, max(len(data.ratios.table.columns) + 1, 6))


def _write_ratio_block(
    sheet: Worksheet,
    start_row: int,
    title: str,
    table: pd.DataFrame,
) -> int:
    sheet.cell(start_row, 1, title)
    sheet.cell(start_row, 1).fill = PatternFill("solid", fgColor=PRIMARY_BLUE)
    sheet.cell(start_row, 1).font = Font(bold=True, color=WHITE)
    _table_header(sheet, start_row + 1, ["Ratio", *table.columns])
    for row, (label, values) in enumerate(table.iterrows(), start_row + 2):
        sheet.cell(row, 1, _display_label(str(label)))
        for column, value in enumerate(values, 2):
            cell = sheet.cell(row, column, _excel_value(value))
            cell.number_format = _ratio_format(str(label))
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
    return start_row + len(table.index) + 1


def _write_dcf_sheet(sheet: Worksheet, data: ModelResultData) -> dict[str, str]:
    years = list(data.dcf.forecast.index)
    _title(sheet, "Discounted Cash Flow Valuation", len(years) + 1)
    sheet["A2"] = "USD millions except per-share values | End-of-year discounting"
    sheet["A2"].font = Font(italic=True, color=PRIMARY_BLUE)
    metrics = [
        ("Terminal method", str(data.dcf.terminal_method)),
        ("Cost of equity", data.dcf.cost_of_equity),
        ("WACC", data.dcf.wacc),
        ("PV of forecast FCF", data.dcf.pv_forecast_fcf),
        ("Terminal value", data.dcf.terminal_value),
        ("PV of terminal value", data.dcf.pv_terminal_value),
        ("Enterprise value", data.dcf.enterprise_value),
        ("Equity value", data.dcf.equity_value),
        ("Implied share price", data.dcf.implied_share_price),
    ]
    _table_header(sheet, 4, ["Valuation metric", "Value"])
    cells: dict[str, str] = {}
    for row, (label, value) in enumerate(metrics, 5):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, _excel_value(value))
        key = label.lower().replace(" ", "_")
        cells[key] = f"'DCF'!$B${row}"
        if "rate" in label.lower() or label in {"WACC", "Cost of equity"}:
            sheet.cell(row, 2).number_format = PERCENT_FORMAT
        elif "price" in label.lower():
            sheet.cell(row, 2).number_format = PRICE_FORMAT
        else:
            sheet.cell(row, 2).number_format = FINANCIAL_FORMAT
    shares = data.dcf.equity_value / data.dcf.implied_share_price
    shares_row = 14
    sheet.cell(shares_row, 1, "Diluted shares (millions)")
    sheet.cell(shares_row, 2, shares)
    sheet.cell(shares_row, 2).number_format = FINANCIAL_FORMAT
    cells["diluted_shares"] = f"'DCF'!$B${shares_row}"

    forecast_header = 17
    _table_header(sheet, forecast_header, ["DCF component", *years])
    dcf_frame = data.dcf.forecast.T
    row_map: dict[str, int] = {}
    for row, (label, values) in enumerate(dcf_frame.iterrows(), forecast_header + 1):
        row_map[str(label)] = row
        sheet.cell(row, 1, _display_label(str(label)))
        for column, value in enumerate(values, 2):
            cell = sheet.cell(row, column, _excel_value(value))
            cell.number_format = _number_format_for(str(label), value)
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
    cells["ufcf_range"] = (
        f"'DCF'!$B${row_map['unlevered_fcf']}:${get_column_letter(len(years) + 1)}"
        f"${row_map['unlevered_fcf']}"
    )
    cells["period_range"] = (
        f"'DCF'!$B${row_map['discount_period']}:${get_column_letter(len(years) + 1)}"
        f"${row_map['discount_period']}"
    )
    cells["last_ufcf"] = f"'DCF'!${get_column_letter(len(years) + 1)}${row_map['unlevered_fcf']}"
    cells["last_period"] = f"'DCF'!${get_column_letter(len(years) + 1)}${row_map['discount_period']}"

    bridge_row = forecast_header + len(dcf_frame.index) + 3
    _table_header(sheet, bridge_row, ["Equity bridge", "Value"])
    adjustment_rows: list[int] = []
    for row, (label, value) in enumerate(data.dcf.equity_bridge.items(), bridge_row + 1):
        sheet.cell(row, 1, _display_label(str(label)))
        sheet.cell(row, 2, float(value))
        sheet.cell(row, 2).number_format = FINANCIAL_FORMAT
        if label != "enterprise_value":
            adjustment_rows.append(row)
    cells["bridge_adjustments"] = "+".join(f"'DCF'!$B${row}" for row in adjustment_rows)
    _size_financial_sheet(sheet, len(years) + 1)
    return cells


def _write_sensitivity_sheet(
    sheet: Worksheet,
    data: ModelResultData,
    dcf_cells: dict[str, str],
) -> None:
    growth_values = list(data.sensitivity.columns)
    wacc_values = list(data.sensitivity.index)
    _title(sheet, "DCF Sensitivity — WACC vs Terminal Growth", len(growth_values) + 1)
    sheet["A2"] = "Formula-driven implied share price; invalid WACC ≤ growth combinations return N/A."
    sheet["A2"].font = Font(italic=True, color=PRIMARY_BLUE)
    _table_header(sheet, 4, ["WACC / Terminal Growth", *growth_values])
    for column in range(2, len(growth_values) + 2):
        sheet.cell(4, column).number_format = PERCENT_FORMAT
    for row, wacc in enumerate(wacc_values, 5):
        sheet.cell(row, 1, float(wacc))
        sheet.cell(row, 1).number_format = PERCENT_FORMAT
        sheet.cell(row, 1).font = Font(color=INPUT_BLUE)
        for column in range(2, len(growth_values) + 2):
            growth_ref = f"{get_column_letter(column)}$4"
            wacc_ref = f"$A{row}"
            formula = (
                f'=IF({wacc_ref}<={growth_ref},NA(),('
                f'SUMPRODUCT({dcf_cells["ufcf_range"]},1/(1+{wacc_ref})^{dcf_cells["period_range"]})+'
                f'{dcf_cells["last_ufcf"]}*(1+{growth_ref})/({wacc_ref}-{growth_ref})/'
                f'(1+{wacc_ref})^{dcf_cells["last_period"]}+'
                f'{dcf_cells["bridge_adjustments"]})/{dcf_cells["diluted_shares"]})'
            )
            cell = sheet.cell(row, column, formula)
            cell.number_format = PRICE_FORMAT
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
    start = "B5"
    end = f"{get_column_letter(len(growth_values) + 1)}{len(wacc_values) + 4}"
    sheet.conditional_formatting.add(
        f"{start}:{end}",
        ColorScaleRule(
            start_type="min", start_color="D9EAF7",
            mid_type="percentile", mid_value=50, mid_color="5B9BD5",
            end_type="max", end_color="17365D",
        ),
    )
    sheet.column_dimensions["A"].width = 24
    for column in range(2, len(growth_values) + 2):
        sheet.column_dimensions[get_column_letter(column)].width = 15


def _write_checks_sheet(sheet: Worksheet, data: ModelResultData) -> int:
    _title(sheet, "Model Checks", 9)
    sheet["A2"] = "Every check reports actual, expected, difference, tolerance, status, and guidance."
    sheet["A2"].font = Font(italic=True, color=PRIMARY_BLUE)
    headers = [
        "Check", "Actual", "Expected", "Difference", "Tolerance", "Status",
        "Failure Severity", "Fiscal Year", "Message",
    ]
    _table_header(sheet, 4, headers)
    checks = (*data.historical_checks, *data.forecast.checks, *data.dcf.checks)
    for row, item in enumerate(checks, 5):
        values = [
            item.check,
            item.actual,
            item.expected,
            item.difference,
            item.tolerance,
            item.status.value,
            item.severity.value,
            item.context.get("fiscal_year"),
            item.message or "",
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, _excel_value(value))
            if column in {2, 3, 4, 5}:
                cell.number_format = '0.000000;[Red](0.000000);-'
        sheet.cell(row, 9).alignment = Alignment(wrap_text=True)
    last_row = max(5, 4 + len(checks))
    status_range = f"F5:F{last_row}"
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=LIGHT_BLUE)),
    )
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"WARN"'], fill=PatternFill("solid", fgColor=AMBER)),
    )
    widths = [34, 14, 14, 14, 14, 14, 14, 12, 55]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    return last_row


def _write_summary(
    sheet: Worksheet,
    data: ModelResultData,
    dcf_cells: dict[str, str],
    check_last_row: int,
) -> None:
    _title(sheet, f"{data.company_name} — Financial Model & Valuation", 14)
    sheet.freeze_panes = "A4"
    metadata = [
        ("Ticker", data.ticker),
        ("As of", data.as_of),
        ("Units", "USD millions except per-share values"),
        ("Framework", "Modular, explainable, research-oriented, semi-automated"),
    ]
    for row, (label, value) in enumerate(metadata, 3):
        sheet.cell(row, 1, label).font = Font(bold=True, color=PRIMARY_BLUE)
        sheet.cell(row, 2, value)
    sheet["A8"] = "Valuation Summary"
    sheet["A8"].fill = PatternFill("solid", fgColor=PRIMARY_BLUE)
    sheet["A8"].font = Font(bold=True, color=WHITE)
    summary = [
        ("WACC", f"={dcf_cells['wacc']}", PERCENT_FORMAT),
        ("Enterprise Value", f"={dcf_cells['enterprise_value']}", FINANCIAL_FORMAT),
        ("Equity Value", f"={dcf_cells['equity_value']}", FINANCIAL_FORMAT),
        ("Implied Share Price", f"={dcf_cells['implied_share_price']}", PRICE_FORMAT),
        (
            "Model Status",
            f'=IF(COUNTIF(\'Checks\'!$F$5:$F${check_last_row},"FAIL")>0,"FAIL",'
            f'IF(COUNTIF(\'Checks\'!$F$5:$F${check_last_row},"WARN")>0,"WARN","PASS"))',
            "General",
        ),
    ]
    for row, (label, formula, number_format) in enumerate(summary, 9):
        sheet.cell(row, 1, label)
        cell = sheet.cell(row, 2, formula)
        cell.font = Font(bold=True, color=GREEN)
        cell.number_format = number_format
        cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
    sheet.conditional_formatting.add(
        "B13", CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor="F4CCCC"))
    )
    sheet.conditional_formatting.add(
        "B13", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=LIGHT_BLUE))
    )

    sheet["A16"] = "Forecast Overview"
    sheet["A16"].fill = PatternFill("solid", fgColor=PRIMARY_BLUE)
    sheet["A16"].font = Font(bold=True, color=WHITE)
    years = list(data.forecast.income_statement.columns)
    _table_header(sheet, 17, ["Metric", *years])
    overview = (
        ("Revenue", data.forecast.income_statement.loc["revenue"]),
        ("EBITDA", data.forecast.income_statement.loc["ebitda"]),
        ("Operating Margin", data.ratios.table.loc["operating_margin"]),
        ("Net Income", data.forecast.income_statement.loc["net_income"]),
        ("Cash", data.forecast.balance_sheet.loc["cash_and_equivalents"]),
        (
            "Debt",
            data.forecast.balance_sheet.loc["short_term_debt"]
            + data.forecast.balance_sheet.loc["long_term_debt"],
        ),
    )
    for row, (label, values) in enumerate(overview, 18):
        sheet.cell(row, 1, label)
        for column, value in enumerate(values, 2):
            cell = sheet.cell(row, column, float(value))
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
            cell.number_format = PERCENT_FORMAT if "Margin" in label else FINANCIAL_FORMAT
    chart = LineChart()
    chart.title = "Revenue and EBITDA Trend"
    chart.style = 13
    chart.height = 7.5
    chart.width = 13.5
    chart.y_axis.title = "USD millions"
    chart.x_axis.title = "Fiscal year"
    data_ref = Reference(sheet, min_col=1, max_col=len(years) + 1, min_row=18, max_row=19)
    categories = Reference(sheet, min_col=2, max_col=len(years) + 1, min_row=17)
    chart.add_data(data_ref, titles_from_data=True, from_rows=True)
    chart.set_categories(categories)
    chart.series[0].graphicalProperties.line.solidFill = NAVY
    chart.series[1].graphicalProperties.line.solidFill = MEDIUM_BLUE
    chart.series[0].tx = SeriesLabel(v="Revenue")
    chart.series[1].tx = SeriesLabel(v="EBITDA")
    sheet.add_chart(chart, "H3")

    note_row = 26
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=14)
    sheet.cell(
        note_row,
        1,
        "For educational and research purposes only. This workbook does not constitute "
        "investment advice. Review assumptions, data quality, model checks, and limitations "
        "before relying on any output.",
    )
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(note_row, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet.cell(note_row, 1).font = Font(color=PRIMARY_BLUE, italic=True)
    sheet.column_dimensions["A"].width = 28
    for column in range(2, 7):
        sheet.column_dimensions[get_column_letter(column)].width = 15


def _write_sources(sheet: Worksheet, data: ModelResultData) -> None:
    _title(sheet, "Sources and Audit Notes", 10)
    headers = ["Item", "Source", "Period / As-of", "Units", "Status", "Notes"]
    _table_header(sheet, 4, headers)
    opening_source = "Standardized SEC history" if data.historical is not None else "User-supplied configuration"
    opening_status = "CALCULATED" if data.historical is not None else "MANUAL"
    records = [
        (
            "Opening financial state",
            opening_source,
            data.as_of,
            "USD millions",
            opening_status,
            (
                "Derived through the disclosed historical opening-state adapter."
                if data.historical is not None
                else "Replace with SEC-standardized history for live-company runs."
            ),
        ),
        (
            "Forecast assumptions",
            "User-supplied configuration",
            "Forecast",
            "Mixed",
            "MANUAL",
            "Centralized on the Assumptions sheet; editable inputs use blue font.",
        ),
        (
            "Valuation output",
            "Python DCF engine",
            data.as_of,
            "USD millions / per share",
            "CALCULATED",
            "End-of-year discounting. See DCF and Checks sheets.",
        ),
    ]
    for row, record in enumerate(records, 5):
        for column, value in enumerate(record, 1):
            sheet.cell(row, column, value)
        sheet.cell(row, 6).alignment = Alignment(wrap_text=True)
    audit_notes = [
        *(('Limitation', item) for item in data.limitations),
        *(('Opening-state warning', item) for item in data.opening_state_warnings),
    ]
    for offset, (note_type, note) in enumerate(audit_notes, 10):
        sheet.cell(offset, 1, note_type)
        sheet.cell(offset, 2, note)
        sheet.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=6)
        sheet.cell(offset, 2).alignment = Alignment(wrap_text=True)
    if data.historical is not None:
        provenance = data.historical.provenance_frame()
        start_row = max(14, 12 + len(audit_notes))
        sheet.cell(start_row, 1, "Field-Level SEC Provenance")
        sheet.cell(start_row, 1).fill = PatternFill("solid", fgColor=PRIMARY_BLUE)
        sheet.cell(start_row, 1).font = Font(bold=True, color=WHITE)
        columns = [
            "statement", "account", "fiscal_year", "value", "source_tag",
            "source_filing", "filing_date", "unit", "confidence", "selection_method",
        ]
        _table_header(sheet, start_row + 1, columns)
        if provenance.empty:
            sheet.cell(start_row + 2, 1, "No field-level provenance records were supplied.")
        else:
            for row_offset, record in enumerate(
                provenance.reindex(columns=columns).to_dict(orient="records"), start_row + 2
            ):
                for column, key in enumerate(columns, 1):
                    sheet.cell(row_offset, column, _excel_value(record[key]))
        for column in range(7, 11):
            sheet.column_dimensions[get_column_letter(column)].width = 18
    widths = [30, 32, 18, 22, 14, 70]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_historical(sheet: Worksheet, data: ModelResultData) -> None:
    _title(sheet, "Historical Financials", 8)
    if data.historical is not None:
        row = 4
        for name, frame in data.historical.statements.items():
            years = list(frame.columns)
            sheet.cell(row, 1, _display_label(name))
            sheet.cell(row, 1).fill = PatternFill("solid", fgColor=PRIMARY_BLUE)
            sheet.cell(row, 1).font = Font(bold=True, color=WHITE)
            _table_header(sheet, row + 1, ["Account", *years])
            for data_row, (account, values) in enumerate(frame.iterrows(), row + 2):
                sheet.cell(data_row, 1, _display_label(str(account)))
                for column, value in enumerate(values, 2):
                    cell = sheet.cell(data_row, column, _excel_value(value))
                    cell.number_format = _number_format_for(str(account), value)
            row += len(frame.index) + 4
        _size_financial_sheet(sheet, max(len(frame.columns) + 1 for frame in data.historical.statements.values()))
        return
    sheet.merge_cells("A4:H7")
    sheet["A4"] = (
        "Historical SEC statements are not available in this manual opening-state result. "
        "The sheet is reserved for the standardized SEC adapter and will never silently "
        "manufacture missing historical values."
    )
    sheet["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A4"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A4"].font = Font(color=PRIMARY_BLUE)
    sheet.column_dimensions["A"].width = 34


def _size_financial_sheet(sheet: Worksheet, last_column: int) -> None:
    sheet.column_dimensions["A"].width = 38
    for column in range(2, last_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 15


def _ordered_record_columns(records: list[dict[str, object]]) -> list[str]:
    if not records:
        return []
    ordered: list[str] = []
    for record in records:
        for key in record:
            if key not in ordered:
                ordered.append(key)
    return ordered


def _excel_value(value: object) -> object:
    if value is None or pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def _display_label(label: str) -> str:
    return label.replace("_", " ").title()


def _is_total_row(label: str) -> bool:
    tokens = (
        "total_", "gross_profit", "ebitda", "operating_income", "net_income",
        "cash_from_", "net_change", "ending_", "enterprise_value", "equity_value",
    )
    return label.startswith("total_") or any(token in label for token in tokens)


def _is_detail_row(label: str) -> bool:
    return not _is_total_row(label)


def _number_format_for(label: str, value: object) -> str:
    lowered = label.lower()
    if isinstance(value, bool):
        return "General"
    if any(token in lowered for token in ("margin", "rate", "growth", "discount_factor")):
        return PERCENT_FORMAT
    if any(token in lowered for token in ("dso", "dio", "dpo", "days", "period", "iterations")):
        return '0.0;[Red](0.0);-'
    return FINANCIAL_FORMAT


def _ratio_format(label: str) -> str:
    if any(token in label for token in ("growth", "margin", "roa", "roe", "roic", "capex_to_revenue")):
        return PERCENT_FORMAT
    if any(token in label for token in ("dso", "dio", "dpo", "cash_conversion_cycle")):
        return '0.0" days";[Red](0.0" days");-'
    return MULTIPLE_FORMAT


def _assumption_number_format(label: str) -> str:
    if any(token in label for token in ("growth", "pct", "rate", "ratio")):
        return PERCENT_FORMAT
    if "days" in label or "life" in label:
        return '0.0;[Red](0.0);-'
    return FINANCIAL_FORMAT
