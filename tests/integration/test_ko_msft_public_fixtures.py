import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fmva import ValuationModel
from fmva.data.statement_builder import HistoricalStatements


@pytest.mark.parametrize(
    ("ticker", "expected_price"),
    (("ko", 67.72), ("msft", 285.43)),
)
def test_public_fixture_runs_complete_offline_model(
    ticker: str,
    expected_price: float,
) -> None:
    fixture = Path(f"data/sample/{ticker}_fy2021_2025_history.json")
    payload = json.loads(fixture.read_text(encoding="utf-8"))
    history = HistoricalStatements.from_dict(payload)
    model = ValuationModel.from_history_json(
        fixture,
        forecast_assumptions_path=f"examples/{ticker}/forecast_assumptions.yaml",
        valuation_assumptions_path=f"examples/{ticker}/valuation_assumptions.yaml",
    )
    result = model.run()

    assert payload["metadata"]["fiscal_years"] == [2021, 2022, 2023, 2024, 2025]
    assert len(history.observations) == 330
    assert not history.quality_issues
    assert list(result.forecast.income_statement.columns) == [2026, 2027, 2028, 2029, 2030]
    assert all(check.status.value == "PASS" for check in result.historical_checks)
    assert all(check.status.value == "PASS" for check in result.forecast.checks)
    assert all(check.status.value == "PASS" for check in result.dcf.checks)
    assert result.dcf.implied_share_price == pytest.approx(expected_price, abs=0.02)


def test_ko_debt_and_trade_payables_use_explicit_fallbacks() -> None:
    history = HistoricalStatements.read_json(
        Path("data/sample/ko_fy2021_2025_history.json")
    )
    latest = {
        item.account: item
        for item in history.observations
        if item.fiscal_year == 2025
    }

    assert latest["accounts_payable"].provenance.source_tag == "AccountsPayableTradeCurrent"
    assert latest["short_term_debt"].provenance.source_tag == "NotesAndLoansPayable"
    assert (
        latest["long_term_debt"].provenance.source_tag
        == "LongTermDebtAndCapitalLeaseObligationsIncludingCurrentMaturities"
    )


def test_msft_segment_model_and_sourced_kpis_flow_to_outputs(tmp_path: Path) -> None:
    model = ValuationModel.from_history_json(
        "data/sample/msft_fy2021_2025_history.json",
        forecast_assumptions_path="examples/msft/forecast_assumptions.yaml",
        valuation_assumptions_path="examples/msft/valuation_assumptions.yaml",
        business_driver_path="examples/msft/business_drivers.yaml",
        business_kpi_history_path="data/sample/msft_business_kpis_fy2023_2025.csv",
    )
    result = model.run()

    assert result.business_kpi_history is not None
    assert result.forecast.business_drivers is not None
    assert all(check.status.value == "PASS" for check in result.historical_checks)
    assert all(check.status.value == "PASS" for check in result.forecast.checks)
    assert result.forecast.income_statement.loc["revenue", 2026] == pytest.approx(
        result.forecast.business_drivers.loc["total_revenue", 2026]
    )
    assert {check.check for check in result.historical_checks} >= {
        "business_kpi_source_completeness",
        "business_kpi_segment_revenue_tie",
        "business_kpi_segment_cogs_tie",
    }

    workbook = result.export_excel(tmp_path / "msft_segment_model.xlsx")
    report = result.export_markdown(tmp_path / "msft_segment_model.md")
    tables = result.export_tables(tmp_path / "tables")
    assert {"Business_History", "Business_Drivers"} <= set(
        load_workbook(workbook, read_only=True).sheetnames
    )
    assert "## Historical Business KPIs" in report.read_text(encoding="utf-8")
    assert "business_kpi_history" in tables


def test_msft_subscriber_arpu_archetype_runs_linked_model() -> None:
    model = ValuationModel.from_history_json(
        "data/sample/msft_fy2021_2025_history.json",
        forecast_assumptions_path="examples/msft/forecast_assumptions.yaml",
        valuation_assumptions_path="examples/msft/valuation_assumptions.yaml",
        business_driver_path="examples/msft/subscriber_business_drivers.yaml",
    )

    result = model.run()
    drivers = result.forecast.business_drivers

    assert drivers is not None
    assert drivers.loc["microsoft_365_consumer_subscribers_millions", 2026] == pytest.approx(
        96.12
    )
    assert result.forecast.income_statement.loc["revenue", 2026] == pytest.approx(
        drivers.loc["total_revenue", 2026]
    )
    assert all(check.status.value == "PASS" for check in result.forecast.checks)
    assert all(check.status.value == "PASS" for check in result.dcf.checks)
